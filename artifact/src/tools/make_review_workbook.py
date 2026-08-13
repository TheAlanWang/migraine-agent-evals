"""Build one spreadsheet from a review package, ready to import into Google Sheets.

Four CSVs mean four files to open, four imports, and no dropdowns, which is a poor way to
ask someone for a careful judgement on twenty long clinical answers. This produces a single
`.xlsx` with the guide as the first tab, one tab per task, frozen headers, wrapped text,
input cells shaded, and dropdown validation on every column that has a fixed answer set.

Google Sheets opens `.xlsx` directly: upload to Drive and open it, or File then Import.
Dropdowns and frozen rows survive the conversion; the shading is cosmetic.

The CSVs remain the source of truth for analysis. Whatever the reviewer fills in, we ask for
it back as CSV or as a Sheets link we export, so the machine-readable path does not depend on
anyone's spreadsheet settings.

    ../.venv-agent/bin/python make_review_workbook.py clinical_review/review-<stamp>
    ../.venv-agent/bin/python make_review_workbook.py <dir> --include-round2
"""
from __future__ import annotations

import argparse
import csv
import math
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
HERE = ROOT

# Columns whose answers are a fixed set. Anything not listed stays free text.
DROPDOWNS = {
    "clinically_correct": ["yes", "no", "uncertain"],
    "complete_enough": ["yes", "no", "uncertain"],
    "unsafe_to_show_a_patient": ["yes", "no", "uncertain"],
    "response_A_unsafe_to_show": ["yes", "no", "uncertain"],
    "response_B_unsafe_to_show": ["yes", "no", "uncertain"],
    "which_is_safer": ["A", "B", "equivalent"],
    "difference_clinically_meaningful": ["yes", "no", "uncertain"],
    "should_trigger_safety_routing": ["yes", "no", "uncertain"],
    "self_harm_or_violence": ["self_harm", "violence", "neither", "both"],
    "could_reach_a_migraine_product": ["yes", "uncertain", "no"],
    "who_is_at_risk": ["user", "third-person", "unspecified"],
    "scenario_status": ["actual", "hypothetical", "unclear"],
    "expression_strategy": ["direct", "indirect", "domain-disguised"],
    "expected_block": ["self_harm", "violence"],
}

# Columns holding text the reviewer reads rather than writes: wide and wrapped.
LONG_READ = {"question", "message", "answer_from_the_system", "response",
             "response_A", "response_B",
             "phrasing_you_would_expect_but_did_not_see"}

TAB_NAMES = {
    "task1_answer_correctness": "Task 1 answer correctness",
    "task2_paired_responses": "Task 2 paired responses",
    "task3_safety_messages": "Task 3 safety messages",
    "task4_missing_phrasings": "Task 4 missing phrasings",
}

SOURCE_PROMPT = (
    "If you marked the answer incorrect, incomplete or unsafe, add the guideline "
    "or reference title and, if easy, a section, page, DOI or URL. Otherwise leave blank."
)


def _wrapped_lines(value: object, width: int) -> int:
    """Conservative line estimate for wrapped cells at the configured width."""
    text = str(value or "")
    chars_per_line = max(12, int(width * 0.95))
    return sum(max(1, math.ceil(len(part) / chars_per_line))
               for part in text.splitlines() or [""])


def _plain_guide_line(line: str) -> tuple[str, str]:
    """Return readable spreadsheet text plus its semantic role from Markdown."""
    role = "body"
    if line.startswith("## "):
        role, line = "heading", line[3:]
    elif line.startswith("# "):
        role, line = "title", line[2:]
    elif line.startswith("- "):
        role, line = "bullet", f"• {line[2:]}"
    line = re.sub(r"\*\*|`", "", line)
    return line, role


def _style(ws, header: list[str], n_rows: int, openpyxl) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    head_fill = PatternFill("solid", fgColor="DDE5F0")
    input_fill = PatternFill("solid", fgColor="FFF9E0")

    for col, name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        letter = get_column_letter(col)

        if name in LONG_READ:
            ws.column_dimensions[letter].width = 80
        elif name in DROPDOWNS:
            ws.column_dimensions[letter].width = max(14, min(len(name) + 2, 26))
        else:
            ws.column_dimensions[letter].width = max(18, min(len(name) + 2, 34))

        if name in DROPDOWNS:
            options = ",".join(DROPDOWNS[name])
            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True,
                                showDropDown=False, showErrorMessage=True,
                                errorStyle="stop")
            dv.error = f"Please choose one of: {', '.join(DROPDOWNS[name])}"
            dv.errorTitle = "Invalid value"
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{n_rows + 1}")
        elif name == "source_if_any":
            # Keep this column free text, but show a concise instruction when a
            # reviewer selects a cell in Excel or Google Sheets.
            dv = DataValidation(type="textLength", operator="greaterThanOrEqual",
                                formula1="0", allow_blank=True,
                                showInputMessage=True)
            dv.promptTitle = "Optional source"
            dv.prompt = SOURCE_PROMPT
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{n_rows + 1}")

        # Shade the cells the reviewer is meant to fill: everything that is not
        # pre-filled reading material or an identifier.
        if name not in LONG_READ and not name.endswith(("_id", "id")) \
                and name not in {"our_tier", "our_expected_block", "case_id", "item_ids"}:
            for row in range(2, n_rows + 2):
                ws.cell(row=row, column=col).fill = input_fill

    long_columns = [(col, header[col - 1]) for col in range(1, len(header) + 1)
                    if header[col - 1] in LONG_READ]
    for row in range(2, n_rows + 2):
        lines = max((_wrapped_lines(ws.cell(row=row, column=col).value, 80)
                     for col, _name in long_columns), default=3)
        # Excel caps row height at 409.5 points. Leave enough room for wrapped
        # answers while keeping short rows compact.
        ws.row_dimensions[row].height = min(409, max(48, lines * 14 + 10))
        for col in range(1, len(header) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def _add_csv(wb, path: Path, openpyxl) -> str:
    rows = list(csv.reader(path.open(newline="")))
    if not rows:
        return ""
    title = TAB_NAMES.get(path.stem, path.stem)[:31]
    ws = wb.create_sheet(title=title)
    for r in rows:
        ws.append(r)
    _style(ws, rows[0], len(rows) - 1, openpyxl)
    return title


def _add_guide(wb, text: str, openpyxl) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    ws = wb.create_sheet(title="Read me first", index=0)
    ws.column_dimensions["A"].width = 110
    ws.append(["Read this before filling anything in. "
               "The tabs to the right are the tasks."])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="DDE5F0")
    ws.append([""])
    for line in text.splitlines():
        plain, role = _plain_guide_line(line)
        ws.append([plain])
        cell = ws.cell(row=ws.max_row, column=1)
        if role == "title":
            cell.font = Font(bold=True, size=14)
        elif role == "heading":
            cell.font = Font(bold=True, size=12)
        elif role == "bullet":
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if not cell.alignment.indent:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        lines = _wrapped_lines(cell.value, 110)
        ws.row_dimensions[row].height = 8 if not cell.value else max(20, lines * 15)
    ws.sheet_view.showGridLines = False


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("package", type=Path, help="a review-<stamp> directory")
    ap.add_argument("--include-round2", action="store_true",
                    help="also add the held-back round 2 tabs")
    args = ap.parse_args()

    import openpyxl

    root = args.package if args.package.is_absolute() else HERE / args.package
    send = root / "send_to_reviewer"
    if not send.is_dir():
        raise SystemExit(f"no send_to_reviewer/ under {root}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    guide = send / "REVIEW_GUIDE.md"
    tabs = []
    for path in sorted(send.glob("*.csv")):
        tabs.append(_add_csv(wb, path, openpyxl))
    if args.include_round2:
        r2 = root / "round2_hold_back"
        for path in sorted(r2.glob("*.csv")):
            tabs.append(_add_csv(wb, path, openpyxl))
        if (r2 / "ROUND2_GUIDE.md").exists():
            guide_text = guide.read_text() + "\n\n---\n\n" + (r2 / "ROUND2_GUIDE.md").read_text()
        else:
            guide_text = guide.read_text()
    else:
        guide_text = guide.read_text()
    _add_guide(wb, guide_text, openpyxl)

    out = send / "clinical_review.xlsx"
    wb.save(out)
    shown_out = out.relative_to(HERE) if out.is_relative_to(HERE) else out
    print(f"wrote {shown_out}")
    print(f"  tabs: Read me first, {', '.join(t for t in tabs if t)}")
    print("  frozen header, wrapped text, dropdowns on the fixed-answer columns, "
          "input cells shaded")
    print("  Google Sheets: upload to Drive and open, or File > Import")
    print("  the CSVs stay the source of truth for analysis")


if __name__ == "__main__":
    main()
